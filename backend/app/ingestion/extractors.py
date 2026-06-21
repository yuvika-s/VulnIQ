"""
Format extractors — Step 1 of the ingestion pipeline.

Every uploaded file, whatever its format, is turned into LLM-readable text
("documents") plus light structural hints. We do NOT try to map fields here.
The job of this module is purely: get the content out of the container format
and into text the Extraction Agent (Role 4) can reason over.

Supported on day one: JSON, XML, CSV, XLSX, PDF.
Each extractor is defensive — a malformed file yields a best-effort text dump
rather than crashing the whole upload.
"""
from __future__ import annotations

import csv
import io
import json
import os
import xml.etree.ElementTree as ET


# --------------------------------------------------------------------------- #
# Each extractor returns a list of "records". A record is a chunk of content
# that plausibly represents ONE finding (or a small group). The Extraction
# Agent then decides how many findings each record actually contains.
# --------------------------------------------------------------------------- #
def detect_format(filename: str, content: bytes) -> str:
    ext = os.path.splitext(filename.lower())[1].lstrip(".")
    if ext in ("json", "xml", "csv", "xlsx", "xls", "pdf"):
        return "xlsx" if ext == "xls" else ext
    # sniff
    head = content[:512].lstrip()
    if head[:1] in (b"{", b"["):
        return "json"
    if head[:5].lower() == b"<?xml" or head[:1] == b"<":
        return "xml"
    if b"%PDF" in content[:8]:
        return "pdf"
    return "csv"  # last resort, treat as delimited text


def extract(filename: str, content: bytes) -> dict:
    """
    Returns {
      "format": str,
      "records": [str, ...],     # LLM-readable chunks, each ~one finding
      "preview": str,            # short human preview for the UI
      "record_count": int,
      "warnings": [str, ...],
    }
    """
    fmt = detect_format(filename, content)
    try:
        if fmt == "json":
            recs, warn = _extract_json(content)
        elif fmt == "xml":
            recs, warn = _extract_xml(content)
        elif fmt == "csv":
            recs, warn = _extract_csv(content)
        elif fmt == "xlsx":
            recs, warn = _extract_xlsx(content)
        elif fmt == "pdf":
            recs, warn = _extract_pdf(content)
        else:
            recs, warn = [content.decode("utf-8", "replace")], ["unknown format"]
    except Exception as e:  # never let one bad file kill the upload
        recs, warn = [content.decode("utf-8", "replace")[:5000]], [f"extractor error: {e}"]

    recs = [r for r in recs if r and r.strip()]
    preview = (recs[0][:400] + "…") if recs else "(no extractable content)"
    return {"format": fmt, "records": recs, "preview": preview,
            "record_count": len(recs), "warnings": warn}


# --------------------------------------------------------------------------- #
# JSON — native tool exports. Walk for arrays of finding-like objects.
# --------------------------------------------------------------------------- #
def _extract_json(content: bytes):
    data = json.loads(content.decode("utf-8", "replace"))
    warnings = []

    # Heuristic: find the largest list of dicts anywhere in the structure —
    # that's almost always the findings array (results, vulnerabilities, etc.)
    best = _find_finding_array(data)
    if best is None:
        # single object or unknown shape — hand the whole thing over as one record
        return [json.dumps(data, indent=2)[:8000]], ["no obvious findings array; passed whole doc"]

    key, arr = best
    if key:
        warnings.append(f"using array at key '{key}' ({len(arr)} items)")
    return [json.dumps(item, indent=2) for item in arr], warnings


def _find_finding_array(data, key_path=""):
    """Recursively locate the most finding-like array of dicts."""
    candidates = []

    def walk(node, path):
        if isinstance(node, list):
            if node and isinstance(node[0], dict):
                # score by how "finding-like" the keys look
                keys = set().union(*[set(d.keys()) for d in node[:5] if isinstance(d, dict)])
                signal = sum(1 for k in keys for hint in
                             ("sever", "vuln", "cve", "cvss", "title", "desc",
                              "finding", "risk", "issue", "name", "result")
                             if hint in k.lower())
                candidates.append((signal, len(node), path, node))
            for i, item in enumerate(node):
                walk(item, f"{path}[{i}]")
        elif isinstance(node, dict):
            for k, v in node.items():
                walk(v, f"{path}.{k}" if path else k)

    walk(data, key_path)
    if not candidates:
        return None
    # prefer highest finding-signal, then longest
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    sig, length, path, arr = candidates[0]
    return (path, arr)


# --------------------------------------------------------------------------- #
# XML — e.g. Nessus .nessus, Burp XML, generic scanner XML
# --------------------------------------------------------------------------- #
def _extract_xml(content: bytes):
    root = ET.fromstring(content.decode("utf-8", "replace"))
    warnings = []

    # Find repeated element tags — the most frequent leaf-ish element is usually
    # the per-finding record (ReportItem, issue, vulnerability, etc.)
    from collections import Counter
    tag_counts = Counter(el.tag for el in root.iter())
    # candidate finding tags: appear many times and have children/attribs
    finding_tags = [t for t, c in tag_counts.items() if c > 1]
    finding_tags.sort(key=lambda t: tag_counts[t], reverse=True)

    chosen = None
    for t in finding_tags:
        sample = root.iter(t).__next__()
        if list(sample) or sample.attrib:  # has substance
            chosen = t
            break

    if not chosen:
        return [ET.tostring(root, encoding="unicode")[:8000]], ["no repeated finding element; passed whole doc"]

    warnings.append(f"using <{chosen}> elements ({tag_counts[chosen]} found)")
    records = []
    for el in root.iter(chosen):
        parts = []
        for k, v in el.attrib.items():
            parts.append(f"{k}: {v}")
        for child in el.iter():
            if child is el:
                continue
            txt = (child.text or "").strip()
            if txt:
                parts.append(f"{child.tag}: {txt}")
        records.append("\n".join(parts))
    return records, warnings


# --------------------------------------------------------------------------- #
# CSV — exported findings. Each row = one record, with header context.
# --------------------------------------------------------------------------- #
def _extract_csv(content: bytes):
    text = content.decode("utf-8", "replace")
    # sniff delimiter, but validate it actually splits the header into >1 field
    delim = ","
    try:
        dialect = csv.Sniffer().sniff(text[:2000], delimiters=",;\t|")
        delim = dialect.delimiter
    except Exception:
        pass
    first_line = text.splitlines()[0] if text.splitlines() else ""
    if first_line.count(delim) == 0:
        # sniff failed — pick whichever common delimiter appears most in header
        delim = max(",;\t|", key=lambda d: first_line.count(d))
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    records = []
    for row in reader:
        parts = [f"{k}: {v}" for k, v in row.items()
                 if k and v and str(v).strip()]
        if parts:
            records.append("\n".join(parts))
    warn = [f"delimiter '{delim}'"] if records else ["no data rows parsed"]
    return records, warn


# --------------------------------------------------------------------------- #
# XLSX — exported findings across one or more sheets.
# --------------------------------------------------------------------------- #
def _extract_xlsx(content: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = []
    warnings = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # find the header row: first row with mostly non-empty string cells
        header_idx = 0
        for i, r in enumerate(rows[:10]):
            non_empty = [c for c in r if c not in (None, "")]
            if len(non_empty) >= max(2, len(r) // 2):
                header_idx = i
                break
        header = [str(c).strip() if c is not None else f"col{j}"
                  for j, c in enumerate(rows[header_idx])]
        for r in rows[header_idx + 1:]:
            parts = []
            for h, v in zip(header, r):
                if v not in (None, ""):
                    parts.append(f"{h}: {v}")
            if parts:
                records.append(f"[sheet: {ws.title}]\n" + "\n".join(parts))
        warnings.append(f"sheet '{ws.title}': header row {header_idx+1}, "
                        f"{len(rows)-header_idx-1} data rows")
    return records, warnings


# --------------------------------------------------------------------------- #
# PDF — scanner reports, pentest reports. Extract text per page; the agent
# segments findings from prose.
# --------------------------------------------------------------------------- #
def _extract_pdf(content: bytes):
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    warnings = []
    pages = []
    for i, page in enumerate(reader.pages):
        txt = page.extract_text() or ""
        if txt.strip():
            pages.append(txt)
    if not pages:
        return [], ["PDF had no extractable text (scanned image PDF? OCR needed in Phase 2)"]

    full = "\n".join(pages)
    warnings.append(f"{len(pages)} pages, {len(full)} chars")

    # PDFs are prose-heavy. We chunk into ~3500-char windows with overlap so the
    # agent can segment findings without losing context across boundaries.
    CHUNK, OVERLAP = 3500, 400
    chunks = []
    start = 0
    while start < len(full):
        chunks.append(full[start:start + CHUNK])
        start += CHUNK - OVERLAP
    return chunks, warnings
